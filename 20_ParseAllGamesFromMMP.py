import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def parse_data(directory):
    data = []
    needs_parsing = []
    unique_games = set()  # Set to store unique games
    
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            relative_path = os.path.relpath(file_path, directory)
            folder_name = os.path.basename(relative_path)
            
            # Count the number of backslashes in the file path
            backslash_count = relative_path.count('\\')
            
            if backslash_count > 1:
                # Split the relative path using backslashes
                parts = relative_path.split('\\')
                
                if "CPS1+2" in parts:
                    # Get the index of "CPS1+2" in the parts list
                    index = parts.index("CPS1+2")
                    # Extract the console and game
                    console = parts[1]
                    game = parts[-1]
                
                    data.append([console, game])
                elif "CPS3" in parts:
                    # Get the index of "CPS3" in the parts list
                    index = parts.index("CPS3")
                    # Extract the console and game
                    console = parts[1]
                    game = parts[-1]
                
                    data.append([console, game])
                elif "IGS" in parts:
                    # Get the index of "IGS" in the parts list
                    index = parts.index("IGS")
                    # Extract the console and game
                    console = parts[1]
                    game = parts[-1]
                
                    data.append([console, game])
                elif "EASYRPG" in parts:
                    # Get the index of "IGS" in the parts list
                    index = parts.index("EASYRPG")
                    # Extract the console and game
                    console = parts[0]
                    game = parts[1]
                    if game not in unique_games:
                        unique_games.add(game)
                        data.append([console, game])
                elif "GBC" in parts:
                    # Check if ".jpg" is present in any part of the path
                    if any(".jpg" in part for part in parts):
                        continue
                    # Get the index of "GBC" in the parts list
                    index = parts.index("GBC")
                    # Extract the console and game
                    console = parts[1]
                    game = parts[-1]
                    
                    data.append([console, game])
                elif "PS" in parts:
                    # Get the index of "PS" in the parts list
                    index = parts.index("PS")
                    # Extract the console and game
                    console = parts[0]
                    game = parts[1]
                    if game not in unique_games:
                        unique_games.add(game)
                        data.append([console, game])
                else:
                    # File path needs further parsing
                    needs_parsing.append(relative_path)
                    
            else:
                # Replace backslash with forward slash and split the file name into column and game parts
                parts = relative_path.replace('\\', '/').split('/')
                # Splitting the file path based on the last occurrence of '/'
                column = '/'.join(parts[:-1])
                game = parts[-1]
                data.append([column, game])
        # Create DataFrame from the parsed data
    df = pd.DataFrame(data, columns=['Console', 'Game'])
    return df, needs_parsing


# Provide the directory path
directory = 'G:\\Roms'

# Parse the data
parsed_data, paths_needing_parsing = parse_data(directory)

# Create a new Excel workbook
workbook = Workbook()

# Create separate sheets for each console
consoles = parsed_data['Console'].unique()
for console in consoles:
    # Filter the data for the current console
    console_data = parsed_data[parsed_data['Console'] == console]

    # Create a new sheet for the console
    sheet = workbook.create_sheet(title=console)

    # Write the data to the sheet
    for row in dataframe_to_rows(console_data, index=False, header=True):
        sheet.append(row)

# Check if the default sheet exists and remove it
if 'Sheet' in workbook.sheetnames:
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

# Save the workbook as an Excel file
output_file = 'parsed_data.xlsx'
workbook.save(output_file)

print("Parsed data saved as", output_file)